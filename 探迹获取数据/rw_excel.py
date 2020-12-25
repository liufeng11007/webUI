# -*- coding: UTF-8 -*-
# @time     : 2020-12-08 17:40
# @Auther   : Aaron
# @File     : rw_excel.py


from openpyxl import load_workbook


class Testcase:
    pass


class RW_Excel:

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def write_data(self, column, get_data):
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        m_row = ws.max_row
        # m_column = ws.max_column

        ws.cell(m_row+1, column, get_data)
        wb.save(self.filename)

    def write_total(self, row, column, total):
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        ws.cell(row, column, total)
        wb.save(self.filename)

    def read_data(self):

        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        testcase_list = []
        header_list = []

        for row in range(1,ws.max_row+1):
            one_testcase = Testcase()
            for column in range(1, ws.max_column+1):
                one_cell_value = ws.cell(row, column).value
                if row == 1:
                    header_list.append(str(one_cell_value))
                else:
                    key = header_list[column-1]
                    if key == "total":
                        setattr(one_testcase, "total_column", column)
                    elif key == "jxs_total":
                        setattr(one_testcase, "jxs_column", column)
                    else:
                        setattr(one_testcase, key, one_cell_value)

            if row !=1:
                setattr(one_testcase, "row", row)
                testcase_list.append(one_testcase)

        return testcase_list


if __name__ == '__main__':
    filename = r"F:\工作\竞品测试\get_zzzs.xlsx"
    sheetname = "jzy"
    read_test = RW_Excel(filename,sheetname)
    data1 = read_test.read_data()
    print(type(data1[0]))
    data2 = data1[0]
    print(data2.hasMobile)
    print(data2.row)
    print(data2.total)






